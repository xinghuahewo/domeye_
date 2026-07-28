"""按文件内容而不是扩展名读取 INFO Excel 文件。"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterator, Tuple


_OLE2_SIGNATURE = bytes.fromhex("d0cf11e0a1b11ae1")
_OOXML_SIGNATURES = (
    b"PK\x03\x04",
    b"PK\x05\x06",
    b"PK\x07\x08",
)


class ExcelReadError(ValueError):
    """Excel 容器无法识别或读取。"""


def detect_excel_container(path: Path) -> str:
    """返回 ``ooxml`` 或 ``ole2``，禁止根据文件后缀猜测格式。"""

    with path.open("rb") as stream:
        signature = stream.read(8)
    if signature.startswith(_OOXML_SIGNATURES):
        return "ooxml"
    if signature == _OLE2_SIGNATURE:
        return "ole2"
    raise ExcelReadError(
        f"{path.name} Excel 魔数未知：{signature.hex() or '<empty>'}"
    )


def iter_first_sheet_values(path: Path) -> Iterator[Tuple[Any, ...]]:
    """以只读方式流式返回第一个工作表的行值。"""

    container = detect_excel_container(path)
    if container == "ooxml":
        try:
            from openpyxl import load_workbook

            # 传入二进制流，避免 openpyxl 因历史错误的 .xls 后缀拒绝
            # 实际为 OOXML 的文件。
            with path.open("rb") as stream:
                workbook = load_workbook(stream, read_only=True, data_only=True)
                try:
                    worksheet = workbook.worksheets[0]
                    yield from worksheet.iter_rows(values_only=True)
                finally:
                    workbook.close()
        except ExcelReadError:
            raise
        except Exception as exc:
            raise ExcelReadError(
                f"{path.name} OOXML 只读解析失败：{exc}"
            ) from exc
        return

    try:
        import xlrd

        workbook = xlrd.open_workbook(str(path), on_demand=True)
        try:
            worksheet = workbook.sheet_by_index(0)
            for row_index in range(worksheet.nrows):
                yield tuple(worksheet.row_values(row_index))
        finally:
            workbook.release_resources()
    except ExcelReadError:
        raise
    except Exception as exc:
        raise ExcelReadError(f"{path.name} OLE2 XLS 只读解析失败：{exc}") from exc
